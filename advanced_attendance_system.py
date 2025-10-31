"""
ADVANCED FACE RECOGNITION ATTENDANCE SYSTEM WITH FACENET
Complete Professional Implementation with All Features
Uses FaceNet (DeepFace) for high-accuracy face recognition
"""

import cv2
import numpy as np
import os
from datetime import datetime, date
import pickle
import csv
from tkinter import *
from tkinter import ttk, messagebox, filedialog
from PIL import Image, ImageTk
import pandas as pd
import time
from deepface import DeepFace

# ==================== Robust Camera Utilities (Windows-friendly) ====================
def _fourcc_str(value: float) -> str:
    try:
        v = int(value)
        return "".join([chr((v >> 8 * i) & 0xFF) for i in range(4)])
    except Exception:
        return ""  # unknown

def try_open_camera(index: int,
                    backend: int,
                    codec: str | None,
                    width: int,
                    height: int,
                    fps: int,
                    warmup_frames: int) -> tuple[cv2.VideoCapture | None, dict | None]:
    """Attempt to open a camera with specific backend/codec/resolution and ensure frames aren't black.
    Returns (cap, meta) on success, (None, None) on failure.
    meta contains backend_name, codec, width, height, fps, brightness.
    """
    backend_names = {
        getattr(cv2, 'CAP_MSMF', -1): 'MSMF',
        getattr(cv2, 'CAP_DSHOW', -1): 'DSHOW',
        getattr(cv2, 'CAP_ANY', -1): 'ANY'
    }

    cap = cv2.VideoCapture(index, backend)
    if not cap.isOpened():
        return None, None

    # Apply settings
    try:
        cap.set(cv2.CAP_PROP_CONVERT_RGB, 1)
    except Exception:
        pass

    if codec:
        cap.set(cv2.CAP_PROP_FOURCC, cv2.VideoWriter_fourcc(*codec))
    cap.set(cv2.CAP_PROP_FRAME_WIDTH, width)
    cap.set(cv2.CAP_PROP_FRAME_HEIGHT, height)
    cap.set(cv2.CAP_PROP_FPS, fps)

    # Warmup and measure brightness
    ok = 0
    brightness_sum = 0.0
    for _ in range(max(10, warmup_frames)):
        ret, fr = cap.read()
        if not ret or fr is None or fr.size == 0:
            continue
        ok += 1
        brightness_sum += float(fr.mean())

    if ok == 0:
        cap.release()
        return None, None

    avg_brightness = brightness_sum / ok
    # Heuristic: treat as black if average brightness is near 0
    if avg_brightness < 5.0:
        cap.release()
        return None, None

    meta = {
        'backend': backend_names.get(backend, str(backend)),
        'codec': codec or 'DEFAULT',
        'width': int(cap.get(cv2.CAP_PROP_FRAME_WIDTH)),
        'height': int(cap.get(cv2.CAP_PROP_FRAME_HEIGHT)),
        'fps': int(cap.get(cv2.CAP_PROP_FPS) or 0),
        'brightness': avg_brightness,
    }
    return cap, meta

def open_best_camera(status_cb=None) -> tuple[cv2.VideoCapture | None, dict | None]:
    """Try several backend/codec/resolution combos and return the first non-black feed.
    status_cb: optional callable(str) to report progress.
    """
    def say(msg: str):
        if status_cb:
            status_cb(msg)
        else:
            print(msg)

    msmf = getattr(cv2, 'CAP_MSMF', 1400)
    dshow = getattr(cv2, 'CAP_DSHOW', 700)
    anyb = getattr(cv2, 'CAP_ANY', 0)

    combos = [
        # backend, codec, warmup_frames
        (msmf, None, 60),         # MSMF default
        (msmf, 'YUYV', 60),       # MSMF YUYV
        (msmf, 'MJPG', 40),       # MSMF MJPEG
        (dshow, 'MJPG', 30),      # DSHOW MJPEG (common fix on Windows)
        (dshow, None, 30),        # DSHOW default
        (anyb, None, 30),         # Any
    ]
    resolutions = [(1280, 720), (640, 480)]
    fps_list = [30, 25]

    for backend, codec, warmup in combos:
        for (w, h) in resolutions:
            for fps in fps_list:
                say(f"Trying backend={backend} codec={codec or 'DEFAULT'} {w}x{h}@{fps}...")
                cap, meta = try_open_camera(0, backend, codec, w, h, fps, warmup)
                if cap is not None:
                    say(f"✓ Selected: {meta}")
                    return cap, meta
    say("✗ No working camera combination found (all black/failed)")
    return None, None

class AdvancedFaceAttendanceSystem:
    def __init__(self, root):
        self.root = root
        self.root.title("Advanced Face Recognition Attendance System - FaceNet Powered")
        self.root.geometry("1400x900")  # Increased height from 850 to 900
        self.root.configure(bg='#2C3E50')
        
        # Initialize face recognition components
        self.recognizer = cv2.face.LBPHFaceRecognizer_create()
        self.face_cascade = cv2.CascadeClassifier(cv2.data.haarcascades + 'haarcascade_frontalface_default.xml')
        
        # FaceNet settings
        self.use_facenet = True  # Use FaceNet for recognition
        self.facenet_model = "Facenet"  # High accuracy model
        self.facenet_threshold = 10.0  # Distance threshold for recognition (higher = more lenient)
        
        # File paths
        self.model_file = "face_model.yml"
        self.names_file = "face_names.pkl"
        self.students_file = "students_database.csv"
        self.attendance_folder = "attendance_records"
        self.images_folder = "student_images"
        self.unknown_faces_folder = "unknown_faces"
        
        # Create necessary folders
        for folder in [self.attendance_folder, self.images_folder, self.unknown_faces_folder]:
            if not os.path.exists(folder):
                os.makedirs(folder)
        
        # Variables
        self.student_data = []
        self.student_lookup_by_name = {}
        self.names = []
        self.facenet_encodings = {}  # Store FaceNet encodings {name: encoding}
        self.current_video = None
        self.recognition_active = False
        self._last_unknown_saved_at = 0.0
        self.last_recognition_results = {}  # Cache recognition results {face_id: (name, confidence)}
        
        # Setup UI
        self.setup_ui()
        self.load_student_database()
        
    def setup_ui(self):
        """Create the complete user interface"""
        
        # ==================== HEADER ====================
        header_frame = Frame(self.root, bg='#34495E', height=80)
        header_frame.pack(fill=X)
        
        title_label = Label(header_frame, text="🎓 ADVANCED FACE RECOGNITION ATTENDANCE SYSTEM", 
                           font=("Arial", 24, "bold"), bg='#34495E', fg='white')
        title_label.pack(pady=20)
        
        # ==================== MAIN CONTAINER ====================
        main_container = Frame(self.root, bg='#2C3E50')
        main_container.pack(fill=BOTH, expand=True, padx=10, pady=10)
        
        # LEFT PANEL - Control Panel with Scrollbar
        left_panel_container = Frame(main_container, bg='#34495E', width=400, relief=RIDGE, bd=2)
        left_panel_container.pack(side=LEFT, fill=BOTH, padx=5, pady=5)
        
        # Add canvas and scrollbar for left panel
        left_canvas = Canvas(left_panel_container, bg='#34495E', width=380, highlightthickness=0)
        left_scrollbar = Scrollbar(left_panel_container, orient=VERTICAL, command=left_canvas.yview)
        left_panel = Frame(left_canvas, bg='#34495E')
        
        left_panel.bind(
            "<Configure>",
            lambda e: left_canvas.configure(scrollregion=left_canvas.bbox("all"))
        )
        
        left_canvas.create_window((0, 0), window=left_panel, anchor="nw")
        left_canvas.configure(yscrollcommand=left_scrollbar.set)
        
        left_scrollbar.pack(side=RIGHT, fill=Y)
        left_canvas.pack(side=LEFT, fill=BOTH, expand=True)
        
        # Mouse wheel scrolling
        def _on_mousewheel(event):
            left_canvas.yview_scroll(int(-1*(event.delta/120)), "units")
        left_canvas.bind_all("<MouseWheel>", _on_mousewheel)
        
        # RIGHT PANEL - Display Panel
        right_panel = Frame(main_container, bg='#34495E', relief=RIDGE, bd=2)
        right_panel.pack(side=RIGHT, fill=BOTH, expand=True, padx=5, pady=5)
        
        # ==================== LEFT PANEL CONTENT ====================
        
        # Student Registration Section
        reg_frame = LabelFrame(left_panel, text="📝 Student Registration", 
                              font=("Arial", 12, "bold"), bg='#34495E', fg='white', bd=2)
        reg_frame.pack(fill=X, padx=10, pady=10)
        
        # Student ID
        Label(reg_frame, text="Student ID:", bg='#34495E', fg='white', 
              font=("Arial", 10)).grid(row=0, column=0, sticky=W, padx=10, pady=5)
        self.student_id_entry = Entry(reg_frame, font=("Arial", 10), width=25)
        self.student_id_entry.grid(row=0, column=1, padx=10, pady=5)
        
        # Student Name
        Label(reg_frame, text="Student Name:", bg='#34495E', fg='white', 
              font=("Arial", 10)).grid(row=1, column=0, sticky=W, padx=10, pady=5)
        self.student_name_entry = Entry(reg_frame, font=("Arial", 10), width=25)
        self.student_name_entry.grid(row=1, column=1, padx=10, pady=5)
        
        # Department
        Label(reg_frame, text="Department:", bg='#34495E', fg='white', 
              font=("Arial", 10)).grid(row=2, column=0, sticky=W, padx=10, pady=5)
        self.dept_var = StringVar()
        dept_combo = ttk.Combobox(reg_frame, textvariable=self.dept_var, 
                                 values=["Computer Science", "Electronics", "Mechanical", 
                                        "Civil", "IT", "Electrical"], 
                                 font=("Arial", 10), width=23, state="readonly")
        dept_combo.grid(row=2, column=1, padx=10, pady=5)
        dept_combo.current(0)
        
        # Year
        Label(reg_frame, text="Year:", bg='#34495E', fg='white', 
              font=("Arial", 10)).grid(row=3, column=0, sticky=W, padx=10, pady=5)
        self.year_var = StringVar()
        year_combo = ttk.Combobox(reg_frame, textvariable=self.year_var, 
                                 values=["1st Year", "2nd Year", "3rd Year", "4th Year"], 
                                 font=("Arial", 10), width=23, state="readonly")
        year_combo.grid(row=3, column=1, padx=10, pady=5)
        year_combo.current(0)
        
        # Email
        Label(reg_frame, text="Email:", bg='#34495E', fg='white', 
              font=("Arial", 10)).grid(row=4, column=0, sticky=W, padx=10, pady=5)
        self.email_entry = Entry(reg_frame, font=("Arial", 10), width=25)
        self.email_entry.grid(row=4, column=1, padx=10, pady=5)
        
        # Phone
        Label(reg_frame, text="Phone:", bg='#34495E', fg='white', 
              font=("Arial", 10)).grid(row=5, column=0, sticky=W, padx=10, pady=5)
        self.phone_entry = Entry(reg_frame, font=("Arial", 10), width=25)
        self.phone_entry.grid(row=5, column=1, padx=10, pady=5)
        
        # Registration Buttons
        btn_frame = Frame(reg_frame, bg='#34495E')
        btn_frame.grid(row=6, column=0, columnspan=2, pady=10)
        
        Button(btn_frame, text="📷 Capture Photo", command=self.capture_student_photo,
               bg='#27AE60', fg='white', font=("Arial", 10, "bold"), 
               width=15, cursor="hand2").pack(side=LEFT, padx=5)
        
        Button(btn_frame, text="💾 Save Student", command=self.save_student,
               bg='#2980B9', fg='white', font=("Arial", 10, "bold"), 
               width=15, cursor="hand2").pack(side=LEFT, padx=5)
        
        # ==================== THRESHOLD CONTROL SECTION ====================
        threshold_frame = LabelFrame(left_panel, text="🎯 Recognition Threshold Control", 
                                    font=("Arial", 12, "bold"), bg='#34495E', fg='white', bd=2)
        threshold_frame.pack(fill=X, padx=10, pady=10)
        
        # Current threshold display
        Label(threshold_frame, text="Current Match Threshold:", bg='#34495E', fg='white', 
              font=("Arial", 11, "bold")).pack(pady=(10, 5))
        
        self.threshold_display = Label(threshold_frame, text=f"{int(self.facenet_threshold*100)}%", 
                                      bg='#2ECC71', fg='white', font=("Arial", 24, "bold"),
                                      relief=RIDGE, bd=3, padx=20, pady=10)
        self.threshold_display.pack(pady=5)
        
        Label(threshold_frame, text="(Faces must match ≥ this % to mark attendance)", 
              bg='#34495E', fg='#ECF0F1', font=("Arial", 9, "italic")).pack(pady=(0, 10))
        
        # Threshold adjustment controls
        adjust_frame = Frame(threshold_frame, bg='#34495E')
        adjust_frame.pack(fill=X, padx=20, pady=10)
        
        Label(adjust_frame, text="Adjust Threshold:", bg='#34495E', fg='white', 
              font=("Arial", 10, "bold")).grid(row=0, column=0, sticky=W, pady=5)
        
        self.threshold_var = StringVar(value="85")
        threshold_entry = Entry(adjust_frame, textvariable=self.threshold_var, 
                               font=("Arial", 12, "bold"), width=8, justify=CENTER,
                               bg='white', fg='black', relief=RIDGE, bd=2)
        threshold_entry.grid(row=0, column=1, padx=10, pady=5)
        
        Label(adjust_frame, text="%", bg='#34495E', fg='white', 
              font=("Arial", 12, "bold")).grid(row=0, column=2, pady=5)
        
        def update_threshold():
            try:
                val = float(self.threshold_var.get())
                if 50 <= val <= 100:
                    self.facenet_threshold = val / 100.0
                    self.threshold_display.config(text=f"{val:.0f}%")
                    
                    # Update color based on threshold level
                    if val >= 85:
                        self.threshold_display.config(bg='#2ECC71')  # Green - High security
                    elif val >= 70:
                        self.threshold_display.config(bg='#F39C12')  # Orange - Medium
                    else:
                        self.threshold_display.config(bg='#E74C3C')  # Red - Low security
                    
                    self.update_info(f"✓ Threshold updated to {val}% ({self.facenet_threshold:.2f})")
                    messagebox.showinfo("✓ Threshold Updated", 
                        f"Match threshold set to {val}%\n\n✓ Only faces matching ≥{val}% will mark attendance\n\n" +
                        ("✅ HIGH SECURITY" if val >= 85 else "⚠️ MEDIUM SECURITY" if val >= 70 else "❌ LOW SECURITY"))
                else:
                    messagebox.showerror("❌ Error", "Threshold must be between 50% and 100%")
            except:
                messagebox.showerror("❌ Error", "Please enter a valid number")
        
        Button(adjust_frame, text="✓ Update Threshold", command=update_threshold,
               bg='#3498DB', fg='white', font=("Arial", 10, "bold"), 
               cursor="hand2", width=18).grid(row=0, column=3, padx=10, pady=5)
        
        # Quick preset buttons
        preset_frame = Frame(threshold_frame, bg='#34495E')
        preset_frame.pack(fill=X, padx=20, pady=(0, 10))
        
        Label(preset_frame, text="Quick Presets:", bg='#34495E', fg='white', 
              font=("Arial", 9)).pack(side=LEFT, padx=5)
        
        def set_preset(value):
            self.threshold_var.set(str(value))
            update_threshold()
        
        Button(preset_frame, text="95% (Very Strict)", command=lambda: set_preset(95),
               bg='#27AE60', fg='white', font=("Arial", 8), cursor="hand2").pack(side=LEFT, padx=2)
        
        Button(preset_frame, text="85% (Recommended)", command=lambda: set_preset(85),
               bg='#2ECC71', fg='white', font=("Arial", 8), cursor="hand2").pack(side=LEFT, padx=2)
        
        Button(preset_frame, text="75% (Moderate)", command=lambda: set_preset(75),
               bg='#F39C12', fg='white', font=("Arial", 8), cursor="hand2").pack(side=LEFT, padx=2)
        
        # ==================== ATTENDANCE SECTION ====================
        attendance_frame = LabelFrame(left_panel, text="📊 Attendance Management", 
                                     font=("Arial", 12, "bold"), bg='#34495E', fg='white', bd=2)
        attendance_frame.pack(fill=X, padx=10, pady=10)
        
        Button(attendance_frame, text="🎥 Start Recognition", 
               command=self.start_recognition,
               bg='#27AE60', fg='white', font=("Arial", 11, "bold"), 
               width=30, height=2, cursor="hand2").pack(pady=5, padx=10)
        
        Button(attendance_frame, text="⏹️ Stop Recognition", 
               command=self.stop_recognition,
               bg='#E74C3C', fg='white', font=("Arial", 11, "bold"), 
               width=30, height=2, cursor="hand2").pack(pady=5, padx=10)
        
        Button(attendance_frame, text="🔄 Train Model", 
               command=self.train_model,
               bg='#8E44AD', fg='white', font=("Arial", 11, "bold"), 
               width=30, height=2, cursor="hand2").pack(pady=5, padx=10)
        
        # ==================== REPORTS SECTION ====================
        reports_frame = LabelFrame(left_panel, text="📈 Reports & Export", 
                                  font=("Arial", 12, "bold"), bg='#34495E', fg='white', bd=2)
        reports_frame.pack(fill=X, padx=10, pady=10)
        
        Button(reports_frame, text="📄 View Today's Attendance", 
               command=self.view_todays_attendance,
               bg='#3498DB', fg='white', font=("Arial", 10, "bold"), 
               width=30, cursor="hand2").pack(pady=3, padx=10)
        
        Button(reports_frame, text="📅 View All Attendance", 
               command=self.view_all_attendance,
               bg='#3498DB', fg='white', font=("Arial", 10, "bold"), 
               width=30, cursor="hand2").pack(pady=3, padx=10)
        
        Button(reports_frame, text="💾 Export to Excel", 
               command=self.export_to_excel,
               bg='#16A085', fg='white', font=("Arial", 10, "bold"), 
               width=30, cursor="hand2").pack(pady=3, padx=10)
        
        Button(reports_frame, text="👥 View All Students", 
               command=self.view_all_students,
               bg='#F39C12', fg='white', font=("Arial", 10, "bold"), 
               width=30, cursor="hand2").pack(pady=3, padx=10)
        
        # ==================== STATUS BAR (outside scrollable area) ====================
        status_frame = Frame(left_panel_container, bg='#34495E')
        status_frame.pack(fill=X, side=BOTTOM, padx=5, pady=5)
        
        self.status_label = Label(status_frame, text="Status: Ready", 
                                 bg='#2ECC71', fg='white', font=("Arial", 10, "bold"),
                                 relief=SUNKEN, anchor=W)
        self.status_label.pack(fill=X)
        
        # ==================== RIGHT PANEL CONTENT ====================
        
        # Video Display
        self.video_label = Label(right_panel, bg='black')
        self.video_label.pack(fill=BOTH, expand=True, padx=10, pady=10)
        
        # Info Display
        info_frame = Frame(right_panel, bg='#34495E', height=150)
        info_frame.pack(fill=X, padx=10, pady=5)
        
        self.info_text = Text(info_frame, height=8, font=("Courier", 10), 
                             bg='#2C3E50', fg='#ECF0F1', relief=RIDGE, bd=2)
        self.info_text.pack(fill=BOTH, expand=True)
        
        scrollbar = Scrollbar(info_frame, command=self.info_text.yview)
        scrollbar.pack(side=RIGHT, fill=Y)
        self.info_text.config(yscrollcommand=scrollbar.set)
        
        self.update_info("System initialized successfully!\nReady to use.")
    
    def update_status(self, message, color='#2ECC71'):
        """Update status bar"""
        self.status_label.config(text=f"Status: {message}", bg=color)
        self.root.update()
    
    def update_info(self, message):
        """Update info display"""
        timestamp = datetime.now().strftime("%H:%M:%S")
        self.info_text.insert(END, f"[{timestamp}] {message}\n")
        self.info_text.see(END)
        self.root.update()
    
    def load_student_database(self):
        """Load students from database"""
        if not os.path.exists(self.students_file):
            # Create database file
            with open(self.students_file, 'w', newline='') as f:
                writer = csv.writer(f)
                writer.writerow(['ID', 'Name', 'Department', 'Year', 'Email', 'Phone', 'Photo', 'Date_Added'])
            self.update_info("Created new student database")
        else:
            with open(self.students_file, 'r') as f:
                reader = csv.DictReader(f)
                self.student_data = list(reader)
                # Build quick lookup by normalized name
                self.student_lookup_by_name = {}
                for row in self.student_data:
                    name_key = str(row.get('Name', '')).strip().lower()
                    if name_key:
                        self.student_lookup_by_name[name_key] = row
            self.update_info(f"Loaded {len(self.student_data)} students from database")
    
    def capture_student_photo(self):
        """Capture student photo using webcam"""
        student_id = self.student_id_entry.get().strip()
        student_name = self.student_name_entry.get().strip()
        
        if not student_id or not student_name:
            messagebox.showerror("Error", "Please enter Student ID and Name first!")
            return
        
        self.update_status("Opening camera for photo capture...", '#F39C12')
        self.update_info("Opening camera...")

        # Try robust camera open (tests multiple backends/codecs)
        cap, meta = open_best_camera(lambda s: self.update_info(s))
        if cap is None:
            messagebox.showerror("Error", "Camera opened but sent only black frames.\nTry closing other apps using the camera and run again.")
            self.update_status("Camera error!", '#E74C3C')
            return

        print("\n" + "="*60)
        print("📷 CAMERA READY")
        print(meta)
        print("="*60)

        # Create window FIRST
        cv2.namedWindow('Capture Student Photo', cv2.WINDOW_NORMAL)
        cv2.namedWindow('Capture Student Photo', cv2.WINDOW_NORMAL)
        cv2.resizeWindow('Capture Student Photo', 960, 720)

        self.update_info("Camera ready - press SPACE to capture")
        print("Press SPACE to capture | Press ESC to cancel\n")
        
        photo_captured = False
        frame_count = 0
        
        while True:
            ret, frame = cap.read()
            if not ret:
                print(f"⚠ Failed to read frame at count {frame_count}")
                messagebox.showerror("Error", "Camera stopped sending frames!")
                break
            
            if frame is None or frame.size == 0:
                print(f"⚠ Empty frame received at count {frame_count}")
                continue
            
            frame_count += 1
            
            # Debug: Print frame info for first 3 frames
            if frame_count <= 3:
                mean_brightness = frame.mean()
                print(f"Frame {frame_count}: shape={frame.shape}, brightness={mean_brightness:.1f}")
                if mean_brightness < 10:
                    print(f"⚠ WARNING: Very dark frame! Check lighting.")
            
            # Keep the color frame for display
            display_frame = frame.copy()
            
            # Detect faces (use grayscale only for detection, not display)
            gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
            faces = self.face_cascade.detectMultiScale(gray, 1.3, 5)
            
            # Draw rectangles around faces on the COLOR frame
            for (x, y, w, h) in faces:
                cv2.rectangle(display_frame, (x, y), (x+w, y+h), (0, 255, 0), 3)
                cv2.putText(display_frame, "Ready to Capture!", (x, y-15), 
                           cv2.FONT_HERSHEY_SIMPLEX, 0.7, (0, 255, 0), 2)
            
            # Instructions with HIGH QUALITY label and camera meta
            cv2.putText(display_frame, "Press SPACE to capture HIGH QUALITY photo", (10, 40), 
                       cv2.FONT_HERSHEY_SIMPLEX, 0.8, (255, 255, 255), 2)
            cv2.putText(display_frame, "Press ESC to cancel", (10, 80), 
                       cv2.FONT_HERSHEY_SIMPLEX, 0.8, (255, 255, 255), 2)
            meta_text = f"{meta.get('backend','?')}/{meta.get('codec','?')} | {frame.shape[1]}x{frame.shape[0]} | FPS~{meta.get('fps',0)}"
            cv2.putText(display_frame, meta_text, 
                       (10, frame.shape[0]-20), 
                       cv2.FONT_HERSHEY_SIMPLEX, 0.6, (0, 255, 0), 2)
            
            # Show the COLOR frame
            cv2.imshow('Capture Student Photo', display_frame)
            
            key = cv2.waitKey(1) & 0xFF
            
            if key == 27:  # ESC
                self.update_info("Photo capture cancelled")
                print("❌ Capture cancelled by user")
                break
            elif key == 32 or key == ord(' '):  # SPACE (check both ways)
                if len(faces) > 0:
                    # Save the COLOR photo with MAXIMUM QUALITY
                    photo_filename = f"{self.images_folder}/{student_id}_{student_name}.jpg"
                    
                    # Save with 100% JPEG quality (no compression)
                    cv2.imwrite(photo_filename, frame, [cv2.IMWRITE_JPEG_QUALITY, 100])
                    
                    print(f"✓ HIGH QUALITY photo saved!")
                    print(f"  File: {photo_filename}")
                    print(f"  Resolution: {frame.shape[1]}x{frame.shape[0]}")
                    print(f"  Quality: 100% (maximum, no compression)")
                    
                    self.update_info(f"High quality photo saved!")
                    
                    # Generate FaceNet encoding
                    if self.use_facenet:
                        self.update_info("Generating FaceNet encoding...")
                        try:
                            embedding = DeepFace.represent(
                                img_path=photo_filename,
                                model_name=self.facenet_model,
                                enforce_detection=False
                            )
                            
                            # Save encoding
                            encoding_file = f"{self.images_folder}/{student_id}_{student_name}_encoding.pkl"
                            with open(encoding_file, 'wb') as f:
                                pickle.dump(embedding[0]['embedding'], f)
                            self.update_info("FaceNet encoding saved!")
                        except Exception as e:
                            self.update_info(f"Warning: Could not generate FaceNet encoding: {str(e)}")
                    
                    messagebox.showinfo("Success", "Photo captured and encoded successfully!")
                    photo_captured = True
                    break
                else:
                    messagebox.showwarning("Warning", "No face detected! Please ensure your face is visible.")
        
        cap.release()
        cv2.destroyAllWindows()
        
        if photo_captured:
            self.update_status("Photo captured successfully!", '#27AE60')
        else:
            self.update_status("Ready", '#2ECC71')
    
    def save_student(self):
        """Save student to database"""
        student_id = self.student_id_entry.get().strip()
        student_name = self.student_name_entry.get().strip()
        dept = self.dept_var.get()
        year = self.year_var.get()
        email = self.email_entry.get().strip()
        phone = self.phone_entry.get().strip()
        
        if not student_id or not student_name:
            messagebox.showerror("Error", "Student ID and Name are required!")
            return
        
        # Check if photo exists
        photo_filename = f"{self.images_folder}/{student_id}_{student_name}.jpg"
        if not os.path.exists(photo_filename):
            result = messagebox.askyesno("Warning", 
                "No photo found for this student. Save anyway?")
            if not result:
                return
            photo_filename = "No Photo"
        
        # Add to database
        with open(self.students_file, 'a', newline='') as f:
            writer = csv.writer(f)
            writer.writerow([student_id, student_name, dept, year, email, phone, 
                           photo_filename, datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
        
        self.update_info(f"Student saved: {student_id} - {student_name}")
        self.update_status("Student saved successfully!", '#27AE60')
        messagebox.showinfo("Success", f"Student {student_name} added successfully!")
        
        # Clear fields
        self.student_id_entry.delete(0, END)
        self.student_name_entry.delete(0, END)
        self.email_entry.delete(0, END)
        self.phone_entry.delete(0, END)
        
        # Reload database
        self.load_student_database()
    
    def train_model(self):
        """Train face recognition model"""
        self.update_status("Training model...", '#8E44AD')
        self.update_info("Starting model training...")
        
        # Get all images
        image_files = [f for f in os.listdir(self.images_folder) 
                      if f.lower().endswith(('.png', '.jpg', '.jpeg'))]
        
        if len(image_files) == 0:
            messagebox.showerror("Error", "No images found! Please add students first.")
            self.update_status("Training failed!", '#E74C3C')
            return
        
        faces_data = []
        labels = []
        names = []
        
        for idx, image_file in enumerate(image_files):
            # Extract name from filename
            name = os.path.splitext(image_file)[0]
            names.append(name)
            
            # Load image
            image_path = os.path.join(self.images_folder, image_file)
            image = cv2.imread(image_path)
            
            if image is None:
                self.update_info(f"Failed to load: {image_file}")
                continue
            
            # Convert to grayscale
            gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
            
            # Detect faces
            faces = self.face_cascade.detectMultiScale(gray, 1.3, 5)
            
            if len(faces) == 0:
                self.update_info(f"No face found in: {image_file}")
                continue
            
            # Use the first face
            (x, y, w, h) = faces[0]
            face_roi = gray[y:y+h, x:x+w]
            face_roi = cv2.resize(face_roi, (200, 200))
            
            faces_data.append(face_roi)
            labels.append(idx)
            
            self.update_info(f"Processed: {name}")
        
        if len(faces_data) == 0:
            messagebox.showerror("Error", "No valid face data found!")
            self.update_status("Training failed!", '#E74C3C')
            return
        
        # Train recognizer
        self.recognizer.train(faces_data, np.array(labels))
        self.recognizer.write(self.model_file)
        
        # Save names
        with open(self.names_file, 'wb') as f:
            pickle.dump(names, f)
        
        self.names = names
        
        self.update_info(f"Training complete! {len(faces_data)} faces trained.")
        self.update_status("Training complete!", '#27AE60')
        messagebox.showinfo("Success", f"Model trained with {len(faces_data)} faces!")
    
    def load_facenet_encodings(self):
        """Load all FaceNet encodings from pickle files"""
        self.facenet_encodings = {}
        
        if not os.path.exists(self.images_folder):
            messagebox.showerror("Error", "Student images folder not found!")
            return False
            
        encoding_files = [f for f in os.listdir(self.images_folder) if f.endswith('_encoding.pkl')]
        
        if len(encoding_files) == 0:
            messagebox.showerror("Error", "No FaceNet encodings found! Please capture student photos first.")
            return False
        
        self.update_info(f"Found {len(encoding_files)} encoding files")
        
        for encoding_file in encoding_files:
            try:
                # Extract student name from filename (format: ID_NAME_encoding.pkl)
                base_name = encoding_file.replace('_encoding.pkl', '')
                parts = base_name.split('_', 1)  # Split only on first underscore
                
                if len(parts) >= 2:
                    student_name = parts[1]  # Second part is the name
                else:
                    student_name = parts[0]  # Fallback to first part
                    
                with open(f"{self.images_folder}/{encoding_file}", 'rb') as f:
                    encoding = pickle.load(f)
                    self.facenet_encodings[student_name] = encoding
                    
                self.update_info(f"✓ Loaded encoding for: {student_name}")
            except Exception as e:
                self.update_info(f"⚠ Could not load {encoding_file}: {str(e)}")
        
        self.update_info(f"Total encodings loaded: {len(self.facenet_encodings)}")
        self.update_info(f"Names: {list(self.facenet_encodings.keys())}")
        return len(self.facenet_encodings) > 0
    
    def start_recognition(self):
        """Start face recognition"""
        if self.use_facenet:
            # Load FaceNet encodings
            self.update_status("Loading FaceNet encodings...", '#3498DB')
            self.update_info("Loading FaceNet encodings...")
            
            if not self.load_facenet_encodings():
                return
            
            self.update_info("✓ FaceNet encodings loaded successfully!")
        else:
            # Load LBPH model
            if not os.path.exists(self.model_file):
                messagebox.showerror("Error", "No trained model found! Please train the model first.")
                return
            
            self.recognizer.read(self.model_file)
            
            if os.path.exists(self.names_file):
                with open(self.names_file, 'rb') as f:
                    self.names = pickle.load(f)
        
        self.update_status("Recognition active...", '#27AE60')
        self.update_info("Starting face recognition...")

        # Use the same robust camera selector for recognition stream
        self.current_video, meta = open_best_camera(lambda s: self.update_info(s))
        if self.current_video is None:
            messagebox.showerror("Error", "Camera opened but sent only black frames. Close other apps using camera and try again.")
            return
        print(f"Recognition camera selected: {meta}")
        self.update_info("Camera ready - recognizing faces...")
        
        self.recognition_active = True
        self.marked_today = set()
        
        self.recognize_faces()
    
    def recognize_faces(self):
        """Recognize faces from video stream"""
        if not self.recognition_active or self.current_video is None:
            return
        
        ret, frame = self.current_video.read()
        
        if not ret:
            self.stop_recognition()
            return
        
        # Convert to grayscale for detection only
        gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
        
        # Detect faces
        faces = self.face_cascade.detectMultiScale(gray, 1.3, 5)
        
        # Process each face
        for (x, y, w, h) in faces:
            name = "Unknown"
            confidence_display = 0.0
            color = (0, 0, 255)  # Red for unknown
            
            if self.use_facenet:
                # FaceNet recognition
                try:
                    # Extract face region from color frame with padding for better recognition
                    padding = 20
                    y1 = max(0, y - padding)
                    y2 = min(frame.shape[0], y + h + padding)
                    x1 = max(0, x - padding)
                    x2 = min(frame.shape[1], x + w + padding)
                    
                    face_img = frame[y1:y2, x1:x2]
                    
                    # Skip if face is too small
                    if face_img.shape[0] < 50 or face_img.shape[1] < 50:
                        name = "Face Too Small"
                        confidence_display = 0.0
                        continue
                    
                    # Save temporarily for FaceNet processing
                    temp_face_path = "temp_face.jpg"
                    cv2.imwrite(temp_face_path, face_img)
                    
                    # Generate FaceNet embedding for detected face
                    embedding = DeepFace.represent(
                        img_path=temp_face_path,
                        model_name=self.facenet_model,
                        enforce_detection=False
                    )
                    
                    detected_encoding = np.array(embedding[0]['embedding'])
                    
                    # Compare with all stored encodings
                    min_distance = float('inf')
                    recognized_name = "Unknown"
                    best_similarity = -1
                    
                    self.update_info(f"🔍 Comparing against {len(self.facenet_encodings)} students...")
                    
                    for student_name, stored_encoding in self.facenet_encodings.items():
                        # Normalize encodings for better comparison
                        stored_enc_array = np.array(stored_encoding)
                        
                        # Calculate both Euclidean distance and Cosine similarity
                        euclidean_distance = np.linalg.norm(detected_encoding - stored_enc_array)
                        
                        # Cosine similarity (better for face recognition)
                        cosine_similarity = np.dot(detected_encoding, stored_enc_array) / (
                            np.linalg.norm(detected_encoding) * np.linalg.norm(stored_enc_array)
                        )
                        
                        # Log comparisons above 30%
                        if cosine_similarity > 0.30:
                            self.update_info(f"   {student_name}: {cosine_similarity*100:.1f}%")
                        
                        # Use cosine similarity (higher is better)
                        if cosine_similarity > best_similarity:
                            best_similarity = cosine_similarity
                            min_distance = euclidean_distance
                            recognized_name = student_name
                    
                    self.update_info(f"🎯 Best match: {recognized_name} at {best_similarity*100:.1f}%")
                    
                    # Check if similarity is above threshold (use self.facenet_threshold = 0.85)
                    # Cosine similarity ranges from -1 to 1, where 1 means identical
                    if best_similarity >= self.facenet_threshold:  # Use configurable threshold (85% by default)
                        name = recognized_name
                        confidence_display = best_similarity * 100  # Show as percentage
                        color = (0, 255, 0)  # Green for recognized
                        
                        self.update_info(f"✅ MATCHED: {name} at {confidence_display:.1f}%")
                        
                        # Mark attendance
                        if name not in self.marked_today:
                            self.mark_attendance(name)
                            self.marked_today.add(name)
                            messagebox.showinfo("✓ Attendance Marked",
                                f"Attendance marked for:\n\n{name}\n\nMatch: {confidence_display:.1f}%\n\n✓ HIGH ACCURACY (≥{self.facenet_threshold*100:.0f}%)")
                    else:
                        name = f"Unknown ({best_similarity*100:.1f}%)"
                        confidence_display = max(0.0, best_similarity * 100)
                        
                        # Log low matches
                        if best_similarity > 0.30:
                            self.update_info(f"⚠️ LOW MATCH: {recognized_name} at {best_similarity*100:.1f}% (Need ≥{self.facenet_threshold*100:.0f}%)")
                        
                        # Save unknown face snapshot with basic throttling (1 image per 2s)
                        now_ts = time.time()
                        if now_ts - self._last_unknown_saved_at > 2.0:
                            self._last_unknown_saved_at = now_ts
                            try:
                                ts = datetime.now().strftime('%Y%m%d_%H%M%S_%f')
                                unknown_img = frame[max(0, y-20):min(frame.shape[0], y+h+20),
                                                    max(0, x-20):min(frame.shape[1], x+w+20)]
                                if unknown_img.size > 0:
                                    save_path = os.path.join(self.unknown_faces_folder, f"unknown_{ts}.jpg")
                                    cv2.imwrite(save_path, unknown_img, [cv2.IMWRITE_JPEG_QUALITY, 95])
                                    self.update_info(f"Saved unknown face: {save_path}")
                            except Exception as e:
                                self.update_info(f"Could not save unknown face: {e}")
                    
                    # Clean up temp file
                    if os.path.exists(temp_face_path):
                        os.remove(temp_face_path)
                        
                except Exception as e:
                    name = "Error"
                    confidence_display = 0.0
                    self.update_info(f"⚠ Recognition error: {str(e)}")
            
            else:
                # LBPH recognition
                face_roi = gray[y:y+h, x:x+w]
                face_roi = cv2.resize(face_roi, (200, 200))
                
                try:
                    label, confidence = self.recognizer.predict(face_roi)
                    
                    if confidence < 70:  # Recognized
                        name = self.names[label] if label < len(self.names) else "Unknown"
                        confidence_display = confidence
                        color = (0, 255, 0)  # Green
                        
                        # Mark attendance
                        if name != "Unknown" and name not in self.marked_today:
                            self.mark_attendance(name)
                            self.marked_today.add(name)
                    else:
                        name = "Unknown"
                        confidence_display = confidence
                        # Save unknown face snapshot with basic throttling
                        now_ts = time.time()
                        if now_ts - self._last_unknown_saved_at > 2.0:
                            self._last_unknown_saved_at = now_ts
                            try:
                                ts = datetime.now().strftime('%Y%m%d_%H%M%S_%f')
                                unknown_img = frame[max(0, y-20):min(frame.shape[0], y+h+20),
                                                    max(0, x-20):min(frame.shape[1], x+w+20)]
                                if unknown_img.size > 0:
                                    save_path = os.path.join(self.unknown_faces_folder, f"unknown_{ts}.jpg")
                                    cv2.imwrite(save_path, unknown_img, [cv2.IMWRITE_JPEG_QUALITY, 95])
                                    self.update_info(f"Saved unknown face: {save_path}")
                            except Exception as e:
                                self.update_info(f"Could not save unknown face: {e}")
                
                except Exception as e:
                    name = "Error"
                    confidence_display = 0.0
            
            # Draw rectangle and label
            cv2.rectangle(frame, (x, y), (x+w, y+h), color, 3)
            cv2.rectangle(frame, (x, y-40), (x+w, y), color, cv2.FILLED)
            
            # Display name and confidence
            cv2.putText(frame, f"{name} ({confidence_display:.1f}%)", (x+6, y-10), 
                       cv2.FONT_HERSHEY_DUPLEX, 0.7, (255, 255, 255), 2)
            
            # Add "PRESENT" badge if attendance was marked
            if name in self.marked_today and name != "Unknown" and "Unknown" not in name:
                badge_y = y + h + 25
                cv2.rectangle(frame, (x, badge_y-25), (x+w, badge_y), (0, 255, 0), cv2.FILLED)
                cv2.putText(frame, "✓ PRESENT", (x+10, badge_y-5), 
                           cv2.FONT_HERSHEY_DUPLEX, 0.8, (255, 255, 255), 2)
        
        # Add info overlay with enhanced details
        recognition_mode = "FaceNet (Deep Learning)" if self.use_facenet else "LBPH"
        overlay_color = (0, 255, 0) if len(faces) > 0 else (255, 255, 255)
        
        # Background box
        cv2.rectangle(frame, (5, 5), (700, 100), (0, 0, 0), cv2.FILLED)
        cv2.rectangle(frame, (5, 5), (700, 100), overlay_color, 2)
        
        cv2.putText(frame, f"Mode: {recognition_mode}", (15, 30), 
                   cv2.FONT_HERSHEY_SIMPLEX, 0.6, overlay_color, 2)
        cv2.putText(frame, f"Faces: {len(faces)} | Attendance: {len(self.marked_today)}", (15, 55), 
                   cv2.FONT_HERSHEY_SIMPLEX, 0.6, overlay_color, 2)
        cv2.putText(frame, f"Threshold: >={self.facenet_threshold*100:.0f}% | Encodings: {len(self.facenet_encodings)}", (15, 80), 
                   cv2.FONT_HERSHEY_SIMPLEX, 0.6, (0, 255, 255), 2)
        
        # Timestamp
        cv2.putText(frame, datetime.now().strftime("%Y-%m-%d %H:%M:%S"), (10, frame.shape[0]-10), 
                   cv2.FONT_HERSHEY_SIMPLEX, 0.6, (255, 255, 255), 1)
        
        # Convert to PhotoImage for tkinter
        frame_rgb = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
        img = Image.fromarray(frame_rgb)
        img = img.resize((800, 600), Image.Resampling.LANCZOS)
        imgtk = ImageTk.PhotoImage(image=img)
        
        self.video_label.imgtk = imgtk
        self.video_label.configure(image=imgtk)
        
        # Continue recognition
        self.root.after(10, self.recognize_faces)
    
    def stop_recognition(self):
        """Stop face recognition"""
        self.recognition_active = False
        
        if self.current_video:
            self.current_video.release()
            self.current_video = None
        
        # Clear video display
        self.video_label.configure(image='')
        self.video_label.imgtk = None
        
        self.update_status("Recognition stopped", '#2ECC71')
        self.update_info("Face recognition stopped")
    
    def mark_attendance(self, name):
        """Mark attendance for a student. Includes ID and Department if available."""
        today = date.today().strftime("%Y-%m-%d")
        attendance_file = f"{self.attendance_folder}/attendance_{today}.csv"
        
        # Create file if doesn't exist
        if not os.path.exists(attendance_file):
            with open(attendance_file, 'w', newline='') as f:
                writer = csv.writer(f)
                writer.writerow(['ID', 'Name', 'Department', 'Date', 'Time', 'Status'])
        
        # Check if already marked
        try:
            with open(attendance_file, 'r') as f:
                reader = csv.reader(f)
                for row in reader:
                    # Name might be col 1 if header extended; find by presence
                    if len(row) > 1 and name == row[1]:
                        return False  # Already marked
        except:
            pass
        
        # Lookup student details by name (case-insensitive)
        sid = ''
        dept = ''
        if isinstance(name, str):
            rec = self.student_lookup_by_name.get(name.strip().lower())
            if rec:
                sid = rec.get('ID', '')
                dept = rec.get('Department', '')

        # Mark attendance
        with open(attendance_file, 'a', newline='') as f:
            writer = csv.writer(f)
            writer.writerow([sid, name, dept, today, datetime.now().strftime("%H:%M:%S"), "Present"])
        
        self.update_info(f"✓ Attendance marked: {name}")
        return True
    
    def view_todays_attendance(self):
        """View today's attendance"""
        today = date.today().strftime("%Y-%m-%d")
        attendance_file = f"{self.attendance_folder}/attendance_{today}.csv"
        
        if not os.path.exists(attendance_file):
            messagebox.showinfo("Info", "No attendance records for today.")
            return
        
        # Create new window
        view_window = Toplevel(self.root)
        view_window.title(f"Attendance - {today}")
        view_window.geometry("800x600")
        
    # Create treeview
        tree_frame = Frame(view_window)
        tree_frame.pack(fill=BOTH, expand=True, padx=10, pady=10)
        
        tree_scroll = Scrollbar(tree_frame)
        tree_scroll.pack(side=RIGHT, fill=Y)
        
        # Read header to determine columns dynamically
        with open(attendance_file, 'r') as f:
            reader = csv.reader(f)
            header = next(reader, ['Name', 'Date', 'Time', 'Status'])

        tree = ttk.Treeview(tree_frame, yscrollcommand=tree_scroll.set,
                           columns=tuple(header), show='headings')
        tree_scroll.config(command=tree.yview)
        
        # Define headings/columns
        for col in header:
            tree.heading(col, text=col)
            # Reasonable default widths
            width = 150
            if col.lower() in ('name',):
                width = 300
            if col.lower() in ('status',):
                width = 100
            tree.column(col, width=width)
        
        tree.pack(fill=BOTH, expand=True)
        
        # Load data
        with open(attendance_file, 'r') as f:
            reader = csv.reader(f)
            next(reader, None)  # Skip header if present
            for row in reader:
                tree.insert('', END, values=row)
    
    def view_all_attendance(self):
        """View all attendance records"""
        attendance_files = [f for f in os.listdir(self.attendance_folder) if f.endswith('.csv')]
        
        if len(attendance_files) == 0:
            messagebox.showinfo("Info", "No attendance records found.")
            return
        
        # Create new window
        view_window = Toplevel(self.root)
        view_window.title("All Attendance Records")
        view_window.geometry("900x600")
        
        # Create notebook for tabs
        notebook = ttk.Notebook(view_window)
        notebook.pack(fill=BOTH, expand=True, padx=10, pady=10)
        
        # Add tab for each date
        for att_file in sorted(attendance_files, reverse=True):
            date_str = att_file.replace('attendance_', '').replace('.csv', '')
            
            tab = Frame(notebook)
            notebook.add(tab, text=date_str)
            
            # Create treeview
            tree_scroll = Scrollbar(tab)
            tree_scroll.pack(side=RIGHT, fill=Y)
            
            # Read header for this file
            with open(f"{self.attendance_folder}/{att_file}", 'r') as hf:
                reader = csv.reader(hf)
                header = next(reader, ['Name', 'Date', 'Time', 'Status'])
            
            tree = ttk.Treeview(tab, yscrollcommand=tree_scroll.set,
                               columns=tuple(header), show='headings')
            tree_scroll.config(command=tree.yview)
            
            for col in header:
                tree.heading(col, text=col)
                width = 150
                if col.lower() in ('name',):
                    width = 300
                if col.lower() in ('status',):
                    width = 100
                tree.column(col, width=width)
            
            tree.pack(fill=BOTH, expand=True)
            
            # Load data
            with open(f"{self.attendance_folder}/{att_file}", 'r') as f:
                reader = csv.reader(f)
                next(reader, None)  # Skip header if present
                for row in reader:
                    tree.insert('', END, values=row)
    
    def view_all_students(self):
        """View all registered students"""
        if len(self.student_data) == 0:
            messagebox.showinfo("Info", "No students registered yet.")
            return
        
        # Create new window
        view_window = Toplevel(self.root)
        view_window.title("All Students")
        view_window.geometry("1000x600")
        
        # Create treeview
        tree_frame = Frame(view_window)
        tree_frame.pack(fill=BOTH, expand=True, padx=10, pady=10)
        
        tree_scroll = Scrollbar(tree_frame)
        tree_scroll.pack(side=RIGHT, fill=Y)
        
        tree = ttk.Treeview(tree_frame, yscrollcommand=tree_scroll.set,
                           columns=('ID', 'Name', 'Department', 'Year', 'Email', 'Phone'), 
                           show='headings')
        tree_scroll.config(command=tree.yview)
        
        # Define headings
        tree.heading('ID', text='Student ID')
        tree.heading('Name', text='Name')
        tree.heading('Department', text='Department')
        tree.heading('Year', text='Year')
        tree.heading('Email', text='Email')
        tree.heading('Phone', text='Phone')
        
        tree.pack(fill=BOTH, expand=True)
        
        # Load data
        with open(self.students_file, 'r') as f:
            reader = csv.reader(f)
            next(reader)  # Skip header
            for row in reader:
                tree.insert('', END, values=row[:6])
    
    def export_to_excel(self):
        """Export attendance to Excel"""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except:
            messagebox.showerror("Error", "Please install openpyxl: pip install openpyxl")
            return
        
        # Ask for save location
        filename = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            initialfile=f"attendance_report_{datetime.now().strftime('%Y%m%d')}.xlsx"
        )
        
        if not filename:
            return
        
        # Create workbook
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        
        # Add sheet for each attendance file
        attendance_files = [f for f in os.listdir(self.attendance_folder) if f.endswith('.csv')]
        
        for att_file in sorted(attendance_files):
            date_str = att_file.replace('attendance_', '').replace('.csv', '')
            ws = wb.create_sheet(title=date_str)
            
            # Load data
            with open(f"{self.attendance_folder}/{att_file}", 'r') as f:
                reader = csv.reader(f)
                for row in reader:
                    ws.append(row)
            
            # Style header
            for cell in ws[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
        
        # Save
        wb.save(filename)
        
        messagebox.showinfo("Success", f"Attendance exported to:\n{filename}")
        self.update_info(f"Exported to: {filename}")

def main():
    root = Tk()
    app = AdvancedFaceAttendanceSystem(root)
    root.mainloop()

if __name__ == "__main__":
    main()
